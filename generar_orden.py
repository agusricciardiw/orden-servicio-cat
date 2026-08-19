# -*- coding: utf-8 -*-
"""
GENERADOR DE LA ORDEN DE SERVICIO DEL CAT
=========================================

Lee las hojas PREORDEN / PREORDEN FINDE del "SISTEMA DE PREORDEN.xlsx",
agrupa los servicios validados y escribe el Word final con un anexo por
zona comunal (o por base operativa).

USO
    python generar_orden.py
    python generar_orden.py --ambito bases
    python generar_orden.py --finde
    python generar_orden.py --vertical           (por defecto va apaisado)
    python generar_orden.py --zona CENTRO        (un solo anexo, para probar)

Todo lo configurable esta en el bloque CONFIG de aca abajo.
"""
import sys, io, re, copy, argparse, datetime
from pathlib import Path
from collections import OrderedDict

import openpyxl
from docx import Document
from docx.oxml.ns import qn
from docx.oxml import OxmlElement

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ==========================================================================
# CONFIG
# ==========================================================================

AQUI = Path(__file__).parent
# Donde se busca la planilla si no se pasa --xlsx. Se toma la mas reciente
# que coincida con el patron: al bajarla de Excel Online varias veces, el
# sistema le agrega "(1)", "(2)", etc., y renombrarla a mano cada semana es
# justo el paso que se olvida.
# Se busca primero en la carpeta "planilla" de este mismo proyecto: es lo mas
# simple de explicar a quien opera -- baja el libro de Excel Online y lo deja
# ahi. Si esa carpeta esta vacia, se prueba con Descargas.
CARPETA_PLANILLA = AQUI / "planilla"
CARPETA_DESCARGAS = Path.home() / "Downloads"
PATRON_XLSX = "SISTEMA DE PREORDEN*.xlsx"
PLANTILLA_SEMANA = AQUI / "plantilla_OS_SEMANA.docx"
PLANTILLA_FINDE = AQUI / "plantilla_OS_FINDE.docx"
DIR_SALIDA = AQUI / "salida"

# Anexo de imagenes: un Word por orden que se completa a mano durante la
# semana. El generador lo pega al final del documento, despues de las tablas,
# y le agrega su entrada en el indice. Si el archivo no existe, se saltea.
INCLUIR_ANEXO_IMAGENES = True
ANEXO_IMAGENES_SEMANA = AQUI / "anexo_imagenes_SEMANA.docx"
ANEXO_IMAGENES_FINDE = AQUI / "anexo_imagenes_FINDE.docx"

# El anexo de imagenes vuelve a hoja vertical (las fotos se ven mejor).
# 'apaisado' lo deja en el mismo formato que las tablas.
ANEXO_IMAGENES_ORIENTACION = 'vertical'

# Los parrafos del anexo de imagenes que contengan esto no se copian: sirven
# para dejar instrucciones dentro del archivo que se completa a mano.
MARCA_INSTRUCCIONES = '{{INSTRUCCIONES}}'

# La orden semanal es UNA sola e incluye las bases operativas primero y
# despues las zonas comunales. La del fin de semana va aparte.
ORDEN_FAMILIAS = ['BASES OPERATIVAS', 'ZONAS COMUNALES']

# Indice de anexos clickeable (cada renglon salta al anexo)
INDICE_CLICKEABLE = True
ANCHO_INDICE = 7200

# No incluir los anexos que quedaron sin ningun servicio validado.
OMITIR_ANEXOS_VACIOS = True

# Las comunas no trabajan los fines de semana: la orden de finde lleva solo
# bases operativas. Si aparece un servicio de finde con base de despliegue
# comunal, el generador avisa, porque es un error de carga.
FINDE_SOLO_BASES = True

# Orientacion de las paginas de anexos ('apaisado' o 'vertical').
# El texto legal queda siempre vertical: se inserta un salto de seccion.
ORIENTACION = 'apaisado'

# Orden de los servicios dentro de cada anexo:
#   'cobertura'      primero los de semana completa, despues por dia puntual
#                    empezando por los lunes  (el que se usa)
#   'base-turno-dia' agrupa por turno y despues por dia
#   'base-dia-hora'  cronologico
#   'base-hora'      por horario
#   'dia-hora'       cronologico puro, mezclando bases
#   'planilla'       respeta el orden de carga del Excel
ORDEN_SERVICIOS = 'cobertura'

# Orden en que se consideran los turnos al ordenar
ORDEN_TURNO = {
    'TM': 0, 'TT': 1, 'TIN': 2, 'TN': 3,
    'FSD S': 0, 'FSD': 0, 'FSD D': 1,
    'FSI S': 2, 'FSI': 2, 'FSI D': 3,
    'FSN D': 4, 'FSN': 4,
}

# Que comunas integra cada zona (sale del indice de anexos del documento real)
ZONAS = OrderedDict([
    ('CENTRO', [1, 3, 4, 5, 6]),
    ('NORTE',  [2, 12, 13, 14, 15]),
    ('SUR',    [7, 8, 9, 10, 11]),
])

# Orden en que salen los anexos de bases operativas
ORDEN_BASES_OPERATIVAS = [
    'ALEJANDRA BERETTA', 'COCHABAMBA', 'OCAMPO', 'VEDIA', 'CINTHIA CHOQUE',
    'ARAOZ DE LAMADRID', 'COUTURE', 'BRD SARMIENTO', 'BRD TACUARI',
]

# Como se escribe cada base DENTRO de la tabla. El titulo del anexo y el
# indice siguen usando el nombre completo: la abreviatura es solo para la
# columna BASE, que se repite en cada fila.
# Las bases de despliegue van como BDC n (Base de Despliegue Comuna n).
ABREVIAR_BASES = True
ABREVIATURA_COMUNA = 'BDC {n}'
ABREVIATURAS_BASE = {
    'ALEJANDRA BERETTA': 'A. Beretta',
    'COCHABAMBA': 'Cocha',
    'OCAMPO': 'Ocampo',
    'VEDIA': 'Vedia',
    'CINTHIA CHOQUE': 'C. Choque',
    'ARAOZ DE LAMADRID': 'A. Lamadrid',
    'COUTURE': 'Couture',
    'BRD SARMIENTO': 'Brd. Sarmiento',
    'BRD TACUARI': 'Brd. Tacuarí',
}

# Cuando TIPO = MISION, la tarea se prefija con MISION en negrita.
# (Lo pide la Guia para la confeccion de la orden de servicio, pag. 7.)
PREFIJO_MISION = True

# Largo maximo de las columnas de texto. None = copiar tal cual.
MAX_DESCRIPCION = 400
MAX_OBSERVACIONES = 400

# 'MAR A VIE' -> 'MAR-MIE-JUE-VIE'.  Si es False se copia literal.
EXPANDIR_RANGO_DE_DIAS = True

# Reescribir la hora a '08:00 A 19:00'. Apagado: la hora sale tal como se
# cargo en la planilla ('8 A 19HS', '7 A 12HS / 12 A 17HS', etc.). Las
# celdas que Excel guardo como hora de verdad se muestran igual como HH:MM,
# porque el valor crudo es una fraccion del dia y no se entiende.
NORMALIZAR_HORA = False

# Repetir los nombres de columna cuando la tabla corta de pagina
REPETIR_ENCABEZADO = True

# Repetir tambien la banda con el nombre del anexo.
# OJO: Word solo repite filas de encabezado si forman un bloque contiguo
# DESDE LA PRIMERA fila. Si esta banda no se repite, tampoco se repiten los
# nombres de columna. Como la banda esta achicada, dejarla prendida cuesta
# menos de una pagina en toda la orden y evita hojas sin encabezado.
REPETIR_TITULO = True

# Que ninguna fila se parta entre dos paginas. Cuesta algun hueco al pie
# cuando la fila siguiente es alta y no entra, pero garantiza que un
# servicio nunca quede cortado a la mitad.
FILA_ENTERA = True

# Exportar tambien a PDF con el nombre que pide la Guia. Requiere Word.
# El numero sale del titulo de la plantilla ("Orden de Servicio N° 2006-2026-O"
# -> ODS_2006-2026-O.pdf), asi no hay que acordarse de cambiarlo en dos
# lugares. Si el titulo no tiene numero, se usan estos de respaldo.
EXPORTAR_PDF = True
NUMERO_ORDEN = '000'
ANIO_ORDEN = '2026'

# Sacar del documento final los comentarios de Word heredados de la plantilla
LIMPIAR_COMENTARIOS = True

# --------------------------------------------------------------------------
# COLUMNAS DE LA TABLA
# (clave, titulo, ancho_vertical, ancho_apaisado, grupo)
# Las que comparten 'grupo' quedan bajo un encabezado combinado.
# Para sacar una columna, comentala; los anchos se reajustan solos.
# --------------------------------------------------------------------------
CAMPOS_SEMANA = [
    ('dia',           'DIA',           1250, 1450, None),
    ('turno',         'TURNO',          950, 1100, None),
    ('tarea',         'TAREA',         1800, 2350, None),
    ('descripcion',   'DESCRIPCIÓN',   1950, 3350, None),
    ('observaciones', 'OBSERVACIONES', 1400, 2200, None),
    ('direccion',     'DIRECCIÓN',     1600, 2300, None),
    ('hora',          'HORA',           850, 1150, None),
    ('ag_0',          'TM',             380,  500, 'AGENTES POR TURNO'),
    ('ag_1',          'TT',             380,  500, 'AGENTES POR TURNO'),
    ('ag_2',          'TIN',            380,  500, 'AGENTES POR TURNO'),
    ('ag_3',          'TN',             380,  500, 'AGENTES POR TURNO'),
    # Angosta a proposito: "A. Beretta" parte en dos renglones, pero como la
    # altura de la fila la manda DESCRIPCION, no cuesta ni una pagina.
    # Ensancharla a 1150 y achicar DESCRIPCION suma 3 paginas al documento.
    ('base',          'BASE',           850,  950, None),
]

CAMPOS_FINDE = [
    ('dia',           'DIA',           1200, 1400, None),
    ('turno',         'TURNO',          950, 1100, None),
    ('tarea',         'TAREA',         1700, 2250, None),
    ('descripcion',   'DESCRIPCIÓN',   1750, 3100, None),
    ('observaciones', 'OBSERVACIONES', 1350, 2200, None),
    ('direccion',     'DIRECCIÓN',     1500, 2100, None),
    ('hora',          'HORA',           820, 1100, None),
    ('ag_0',          'FSD S',          380,  480, 'AGENTES POR TURNO'),
    ('ag_1',          'FSD D',          380,  480, 'AGENTES POR TURNO'),
    ('ag_2',          'FSI S',          380,  480, 'AGENTES POR TURNO'),
    ('ag_3',          'FSI D',          380,  480, 'AGENTES POR TURNO'),
    ('ag_4',          'FSN D',          380,  480, 'AGENTES POR TURNO'),
    ('base',          'BASE',          1000, 1150, None),
]

# Ancho de la hoja A4 en dxa (1440 = 1 pulgada) y margen lateral que se le
# deja a la tabla de anexos. Ninguna impresora imprime hasta el borde: con
# 567 dxa (1 cm) de aire a cada lado la tabla entra entera en cualquiera.
PAGINA = {'vertical': 11906, 'apaisado': 16838}
MARGEN_ANEXO = 567
MARGEN_SECCION = 1440   # el margen real del documento, no se toca

# El banner del encabezado es una imagen de 8,32 pulgadas: en A4 vertical
# cubre la hoja entera, pero en apaisado (11,69") deja un vacio blanco a la
# derecha. Se puede escalar al ancho de la hoja, PERO al agrandarlo crece
# tambien a lo alto y se come una pulgada de tabla por pagina: la orden de
# comunas pasaba de 43 a 66 paginas. Queda apagado a proposito: el que la
# usa pasa cientos de hojas, y esa pulgada vale mas que el vacio blanco.
BANNER_ANCHO_COMPLETO = False

# Margenes de las paginas de anexos. El banner mide 0,98" de alto, asi que
# con 1,11" arriba queda un respiro sin desperdiciar hoja.
MARGEN_SUP_ANEXO = 1600
MARGEN_INF_ANEXO = 1500


def ancho_tabla(orientacion):
    return PAGINA[orientacion] - 2 * MARGEN_ANEXO


def sangria_tabla():
    """Sangria negativa para que la tabla se salga de los margenes del texto
    y llegue hasta MARGEN_ANEXO del borde de la hoja."""
    return -(MARGEN_SECCION - MARGEN_ANEXO)


# --------------------------------------------------------------------------
# PALETAS. Cambia PALETA para elegir cual se usa.
# --------------------------------------------------------------------------
PALETA = 'institucional'

PALETAS = {
    # Toma los azules del banner de la Direccion: la tabla y el encabezado
    # del documento quedan de la misma familia.
    'institucional': dict(
        titulo_fondo='1B3A54', titulo_texto='FFFFFF',
        grupo_fondo='2F6076', grupo_texto='FFFFFF',
        encabezado_fondo='2F6076', encabezado_texto='FFFFFF',
        dato='FFFFFF', zebra='EDF3F6',
        borde='C3D2DA', borde_fuerte='1B3A54'),

    # Mantiene el amarillo de la Direccion pero mas calmo, y baja el gris
    # del encabezado para que no cargue tanta tinta.
    'amarillo': dict(
        titulo_fondo='F2C230', titulo_texto='1A1A1A',
        grupo_fondo='4A4A4A', grupo_texto='FFFFFF',
        encabezado_fondo='4A4A4A', encabezado_texto='FFFFFF',
        dato='FFFFFF', zebra='F6F4EF',
        borde='D2D2D2', borde_fuerte='4A4A4A'),

    # Minima tinta: pensada para imprimir muchas copias en blanco y negro.
    'sobrio': dict(
        titulo_fondo='E7EDF1', titulo_texto='1B3A54',
        grupo_fondo='F0F0F0', grupo_texto='1B3A54',
        encabezado_fondo='F0F0F0', encabezado_texto='1B3A54',
        dato='FFFFFF', zebra='F8F8F8',
        borde='CFD8DD', borde_fuerte='8FA3AE'),

    # La de la orden de servicio original.
    'original': dict(
        titulo_fondo='FFCC02', titulo_texto='000000',
        grupo_fondo='595959', grupo_texto='FFFFFF',
        encabezado_fondo='595959', encabezado_texto='FFFFFF',
        dato='FFFFFF', zebra=None,
        borde='CCCCCC', borde_fuerte='434343'),
}

USAR_ZEBRA = True       # sombreado de filas alternas
FUENTE = 'Roboto'
TAM = 18            # medios puntos -> 9pt. 20 = 10pt si se quiere mas grande
PADDING_CELDA = 30  # aire arriba y abajo dentro de cada celda (dxa)

W = '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}'

# ==========================================================================
# LECTURA DE LA PLANILLA
# ==========================================================================

# --------------------------------------------------------------------------
# Se lee la hoja de CARGA (SEMANA / FINDE), no la hoja PREORDEN.
#
# PREORDEN es el resultado de dos FILTER() derramados, y eso trae dos
# problemas: (1) al bajar el libro, los valores cacheados del derrame llegan
# incompletos -- la columna del ID venia vacia o en #VALUE! y se perdian
# cientos de servicios; (2) el rango del FILTER esta escrito a mano
# (SEMANA!B7:Q482) y se queda corto cuando la hoja crece, dejando afuera todo
# lo que se cargue mas abajo, sin avisar.
#
# Leyendo la hoja de carga y aplicando aca la misma condicion (la casilla
# AÑADIDO) el resultado es identico al que deberia dar PREORDEN, pero no
# depende ni del cache de Excel ni de que alguien mantenga el rango.
# --------------------------------------------------------------------------
HOJA_CARGA = {False: 'SEMANA', True: 'FINDE'}
FILA_ENCABEZADO = 6
COLUMNA_VALIDACION = 'AÑADIDO'
PREFIJO_DOTACION = 'AT '

# Nombre de encabezado -> clave interna
MAPA_COLUMNAS = {
    'ID': 'id', 'TIPO': 'tipo', 'SERVICIO': 'servicio',
    'UBICACION': 'ubicacion', 'UBICACIÓN': 'ubicacion',
    'ALTURA': 'altura', 'CALLE 2': 'calle2', 'CALLE 3': 'calle3',
    'BASE': 'base', 'OBSERVACIONES': 'obs', 'FUNCION': 'funcion',
    'FUNCIÓN': 'funcion', 'DIA': 'dia', 'DÍA': 'dia', 'HORA': 'hora',
    'TURNO': 'turno',
}

DIAS = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'DOM']
ALIAS_DIA = {
    'LUNES': 'LUN', 'LUNE': 'LUN', 'LUN': 'LUN',
    'MARTES': 'MAR', 'MART': 'MAR', 'MAR': 'MAR',
    'MIERCOLES': 'MIE', 'MIÉRCOLES': 'MIE', 'MIER': 'MIE', 'MIÉR': 'MIE',
    'MIE': 'MIE', 'MIÉ': 'MIE', 'MIERC': 'MIE',
    'JUEVES': 'JUE', 'JUEV': 'JUE', 'JUE': 'JUE',
    'VIERNES': 'VIE', 'VIERN': 'VIE', 'VIER': 'VIE', 'VIE': 'VIE',
    'SABADO': 'SAB', 'SÁBADO': 'SAB', 'SAB': 'SAB', 'SÁB': 'SAB',
    'DOMINGO': 'DOM', 'DOM': 'DOM',
}
DIAS_HABILES = {0, 1, 2, 3, 4}   # LUN a VIE


def limpiar(v):
    """FILTRAR() convierte las celdas vacias en 0: hay que sacarlos."""
    if v is None:
        return ''
    if isinstance(v, (int, float)) and v == 0:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return '' if s in ('0', 'None') else s


def fmt_dia(v):
    s = limpiar(v).upper().replace('.', '').strip()
    if not s:
        return ''
    s = re.sub(r'^DE\s+', '', s)          # 'DE LUN A VIE' -> 'LUN A VIE'
    if not EXPANDIR_RANGO_DE_DIAS:
        return s
    m = re.match(r'^\s*([A-ZÁÉÍÓÚ]+)\s+A\s+([A-ZÁÉÍÓÚ]+)\s*$', s)
    if m:
        a, b = ALIAS_DIA.get(m.group(1)), ALIAS_DIA.get(m.group(2))
        if a in DIAS and b in DIAS:
            i, j = DIAS.index(a), DIAS.index(b)
            if i <= j:
                return '-'.join(DIAS[i:j + 1])
    # OJO: no separar por '/', que rompe las fechas tipo 'MAR 18/08'
    partes = [p.strip() for p in re.split(r'[-,]| Y ', s) if p.strip()]
    return '-'.join(ALIAS_DIA.get(p, p) for p in partes)


def dias_de(s):
    """Indices de los dias que cubre un texto de DIA ya formateado.
    Tolera fechas sueltas: de 'MAR 18/08' saca el martes."""
    encontrados = set()
    for tok in re.split(r'[-,\s/]+', (s or '').upper().replace('.', '')):
        d = ALIAS_DIA.get(tok)
        if d:
            encontrados.add(DIAS.index(d))
    return encontrados


def clave_cobertura(s):
    """Primero los servicios que cubren la semana completa; despues los de
    dias puntuales, empezando por los lunes. A igual dia de arranque, va
    antes el que cubre mas dias."""
    d = dias_de(s)
    if d and DIAS_HABILES <= d:
        return (0, 0, 0)
    return (1, min(d) if d else 99, -len(d))


def es_hora_numerica(v):
    """Excel guarda las horas como fraccion del dia: 0.5 = 12:00. Cuando el
    FILTRAR() las derrama a PREORDEN se pierde el formato y llegan asi."""
    return isinstance(v, float) and 0 < v < 1


def fmt_hora(v):
    if isinstance(v, (datetime.time, datetime.datetime)):
        return v.strftime('%H:%M')
    if isinstance(v, datetime.timedelta):
        m = int(v.total_seconds() // 60)
        return f"{m // 60:02d}:{m % 60:02d}"
    if es_hora_numerica(v):
        m = int(round(v * 24 * 60))
        return f"{m // 60:02d}:{m % 60:02d}"
    s = limpiar(v)
    if not s or not NORMALIZAR_HORA:
        return s
    salida = []
    for linea in s.split('\n'):
        linea = linea.strip()
        m = re.match(r'^(\d{1,2})(?::(\d{2}))?\s*(?:A|a)\s*(\d{1,2})(?::(\d{2}))?\s*(?:HS|Hs|hs|H|h)?\.?$', linea)
        if m:
            h1, m1, h2, m2 = m.group(1), m.group(2) or '00', m.group(3), m.group(4) or '00'
            salida.append(f"{int(h1):02d}:{m1} A {int(h2):02d}:{m2}")
        else:
            salida.append(linea)
    return '\n'.join(salida)


def fmt_direccion(ubic, altura, c2, c3):
    ubic, altura, c2, c3 = (limpiar(x) for x in (ubic, altura, c2, c3))
    if not ubic:
        return ''
    if altura:
        return f"-{ubic} {altura}"
    if c2 and c3:
        return f"-{ubic} e/{c2} y {c3}"
    if c2:
        return f"-{ubic} y {c2}"
    return f"-{ubic}"


RE_COMUNA = re.compile(r'DESPLIEGUE\s+COMUNA\s+(\d+)', re.I)


def num_comuna(base):
    m = RE_COMUNA.search(limpiar(base))
    return int(m.group(1)) if m else None


def fmt_base(base):
    """Como se muestra la base en la columna BASE de la tabla."""
    base = limpiar(base)
    n = num_comuna(base)
    if n:
        return ABREVIATURA_COMUNA.format(n=n) if ABREVIAR_BASES \
            else f"BD COMUNA {n}"
    if ABREVIAR_BASES:
        return ABREVIATURAS_BASE.get(base.upper(), base)
    return base


def clave_dia(s):
    for i, d in enumerate(DIAS):
        if s.startswith(d):
            return i
    return 99


def clave_hora(s):
    m = re.search(r'(\d{1,2}):(\d{2})', s or '')
    return int(m.group(1)) * 60 + int(m.group(2)) if m else 9999


def clave_turno(s):
    """TM-TT ordena por su primer turno; lo desconocido va al final."""
    s = (s or '').strip().upper()
    if not s:
        return (99, '')
    primero = re.split(r'[-/]', s)[0].strip()
    return (ORDEN_TURNO.get(primero, ORDEN_TURNO.get(s, 98)), s)


def mapear_columnas(ws):
    """Ubica las columnas por el texto del encabezado, no por posicion fija:
    si alguien inserta una columna en la planilla, esto sigue funcionando."""
    cols, dotacion = {}, []
    validacion = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=FILA_ENCABEZADO, column=c).value
        if v is None:
            continue
        nombre = str(v).strip().upper()
        if nombre in MAPA_COLUMNAS:
            cols.setdefault(MAPA_COLUMNAS[nombre], c)
        elif nombre.startswith(PREFIJO_DOTACION):
            dotacion.append((c, str(v).strip()))
        elif nombre == COLUMNA_VALIDACION:
            validacion = c
    faltan = {'id', 'base', 'servicio', 'funcion', 'dia'} - set(cols)
    if faltan:
        sys.exit(f"En la hoja {ws.title!r} no encuentro las columnas: "
                 f"{', '.join(sorted(faltan))}")
    if validacion is None:
        sys.exit(f"En la hoja {ws.title!r} no encuentro la columna "
                 f"{COLUMNA_VALIDACION!r}, que es la que marca los servicios "
                 f"validados.")
    return cols, dotacion, validacion


def leer_servicios(xlsx, finde):
    hoja = HOJA_CARGA[bool(finde)]
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if hoja not in wb.sheetnames:
        sys.exit(f"La hoja {hoja!r} no existe en {xlsx}")
    ws = wb[hoja]
    COLS, dotacion_cols, col_val = mapear_columnas(ws)

    esperadas = len([c for c in (CAMPOS_FINDE if finde else CAMPOS_SEMANA)
                     if c[0].startswith('ag_')])
    if len(dotacion_cols) != esperadas:
        print(f"  !! La hoja {hoja} tiene {len(dotacion_cols)} columnas de "
              f"dotacion ({', '.join(n for _, n in dotacion_cols)})\n"
              f"     pero la tabla esta definida con {esperadas}. "
              f"Revisar CAMPOS_{'FINDE' if finde else 'SEMANA'}.")

    filas, sin_base, cargados = [], [], 0
    for r in range(7, ws.max_row + 1):
        idv = limpiar(ws.cell(row=r, column=COLS['id']).value)
        if not idv:
            continue
        cargados += 1
        if ws.cell(row=r, column=col_val).value is not True:
            continue          # todavia no validado
        g = lambda k: ws.cell(row=r, column=COLS[k]).value if k in COLS else None

        dot = []
        for c, _ in dotacion_cols:
            v = ws.cell(row=r, column=c).value
            dot.append(int(v) if isinstance(v, (int, float)) else 0)

        base_cruda = limpiar(g('base'))
        tipo = limpiar(g('tipo')).upper()
        servicio = limpiar(g('servicio'))
        es_mision = PREFIJO_MISION and tipo.startswith('MISION') \
            and not servicio.upper().startswith(('MISION', 'MISIÓN'))
        tarea = [('MISIÓN ', True), (servicio, False)] if es_mision \
            else [(servicio, False)]

        desc_completa = limpiar(g('funcion'))
        desc = desc_completa
        recortada = False
        if MAX_DESCRIPCION and len(desc) > MAX_DESCRIPCION:
            desc = desc[:MAX_DESCRIPCION].rstrip() + '…'
            recortada = True

        obs = limpiar(g('obs'))
        if MAX_OBSERVACIONES and len(obs) > MAX_OBSERVACIONES:
            obs = obs[:MAX_OBSERVACIONES].rstrip() + '…'

        reg = dict(
            fila=r, id=idv, base_cruda=base_cruda, tipo=tipo,
            largo_desc=len(desc_completa), recortada=recortada,
            hora_numerica=es_hora_numerica(g('hora')),
            dia=fmt_dia(g('dia')),
            turno=re.sub(r'\s+', ' ', re.sub(r'\s*-\s*', '-',
                                             limpiar(g('turno')))).strip(),
            tarea=tarea,
            descripcion=desc,
            direccion=fmt_direccion(g('ubicacion'), g('altura'),
                                    g('calle2'), g('calle3')),
            hora=fmt_hora(g('hora')),
            observaciones=obs,
            base=fmt_base(base_cruda),
            comuna=num_comuna(base_cruda),
            total_ag=sum(dot),
        )
        for i, v in enumerate(dot):
            reg[f'ag_{i}'] = str(v) if v else ''
        if not base_cruda:
            sin_base.append(reg)
            continue
        filas.append(reg)

    validados = len(filas) + len(sin_base)
    print(f"  Hoja {hoja}: {cargados} servicios cargados, {validados} "
          f"validados ({COLUMNA_VALIDACION} tildado)")
    return filas, sin_base


# ==========================================================================
# CONSTRUCCION DE LAS TABLAS DE WORD
# ==========================================================================

def el(tag, **attrs):
    e = OxmlElement(tag)
    for k, v in attrs.items():
        e.set(qn(k), str(v))
    return e


def bordes(tc_pr, top, left, bottom, right):
    b = el('w:tcBorders')
    for lado, color in (('top', top), ('left', left),
                        ('bottom', bottom), ('right', right)):
        b.append(el(f'w:{lado}', **{'w:val': 'single', 'w:sz': '5',
                                    'w:space': '0', 'w:color': color}))
    tc_pr.append(b)


def celda(texto, ancho, fill, *, negrita=False, color=None, gridspan=None,
          vmerge=None, alineacion='center', enlace=None, subrayado=False,
          tam=None, bordes_=('CCCCCC', 'CCCCCC', 'D9D9D9', 'D9D9D9'),
          margen_lat=120):
    tc = el('w:tc')
    pr = el('w:tcPr')
    pr.append(el('w:tcW', **{'w:w': ancho, 'w:type': 'dxa'}))
    if gridspan:
        pr.append(el('w:gridSpan', **{'w:val': gridspan}))
    if vmerge == 'restart':
        pr.append(el('w:vMerge', **{'w:val': 'restart'}))
    elif vmerge == 'continue':
        pr.append(el('w:vMerge'))
    bordes(pr, *bordes_)
    pr.append(el('w:shd', **{'w:val': 'clear', 'w:color': 'auto', 'w:fill': fill}))
    mar = el('w:tcMar')
    for lado, v in (('top', PADDING_CELDA), ('left', margen_lat),
                    ('bottom', PADDING_CELDA), ('right', margen_lat)):
        mar.append(el(f'w:{lado}', **{'w:w': v, 'w:type': 'dxa'}))
    pr.append(mar)
    pr.append(el('w:vAlign', **{'w:val': 'center'}))
    tc.append(pr)

    p = el('w:p')
    p_pr = el('w:pPr')
    p_pr.append(el('w:widowControl', **{'w:val': '0'}))
    p_pr.append(el('w:jc', **{'w:val': alineacion}))
    p.append(p_pr)

    # Si la celda es un link, las corridas van adentro del w:hyperlink
    contenedor = p
    if enlace:
        contenedor = el('w:hyperlink', **{'w:anchor': enlace})
        p.append(contenedor)

    tramos = texto if isinstance(texto, list) else [(str(texto), negrita)]
    for txt, bold in tramos:
        if txt == '':
            continue
        r = el('w:r')
        r_pr = el('w:rPr')
        r_pr.append(el('w:rFonts', **{'w:ascii': FUENTE, 'w:eastAsia': FUENTE,
                                      'w:hAnsi': FUENTE, 'w:cs': FUENTE}))
        if bold:
            r_pr.append(el('w:b'))
            r_pr.append(el('w:bCs'))
        if color:
            r_pr.append(el('w:color', **{'w:val': color}))
        if subrayado:
            r_pr.append(el('w:u', **{'w:val': 'single'}))
        medida = tam or TAM
        r_pr.append(el('w:sz', **{'w:val': medida}))
        r_pr.append(el('w:szCs', **{'w:val': medida}))
        r.append(r_pr)
        for i, linea in enumerate(str(txt).split('\n')):
            if i:
                r.append(el('w:br'))
            t = el('w:t')
            t.set(qn('xml:space'), 'preserve')
            t.text = linea
            r.append(t)
        contenedor.append(r)
    if not p.findall(f'.//{W}r'):
        p.append(el('w:r'))
    tc.append(p)
    return tc


def fila(celdas, alto=None, encabezado_repetible=False, entera=False):
    tr = el('w:tr')
    pr = el('w:trPr')
    if entera:
        pr.append(el('w:cantSplit'))   # que la fila no se corte entre paginas
    if alto:
        pr.append(el('w:trHeight', **{'w:val': alto}))
    if encabezado_repetible:
        pr.append(el('w:tblHeader'))
    tr.append(pr)
    for c in celdas:
        tr.append(c)
    return tr


def anchos(campos, orientacion):
    """Devuelve los anchos escalados para que sumen el ancho de la tabla."""
    idx = 2 if orientacion == 'vertical' else 3
    crudos = [c[idx] for c in campos]
    total = ancho_tabla(orientacion)
    escala = total / sum(crudos)
    ws = [int(round(w * escala)) for w in crudos]
    ws[-1] += total - sum(ws)   # que cierre exacto
    return ws


def tabla_anexo(titulo, registros, campos, orientacion):
    P = PALETAS[PALETA]
    ws = anchos(campos, orientacion)
    ancho_total = ancho_tabla(orientacion)

    tbl = el('w:tbl')
    pr = el('w:tblPr')
    pr.append(el('w:tblStyle', **{'w:val': 'a0'}))
    pr.append(el('w:tblW', **{'w:w': ancho_total, 'w:type': 'dxa'}))
    pr.append(el('w:tblInd', **{'w:w': sangria_tabla(), 'w:type': 'dxa'}))
    tb = el('w:tblBorders')
    for lado in ('top', 'left', 'bottom', 'right', 'insideH', 'insideV'):
        tb.append(el(f'w:{lado}', **{'w:val': 'nil'}))
    pr.append(tb)
    pr.append(el('w:tblLayout', **{'w:type': 'fixed'}))
    tbl.append(pr)

    grid = el('w:tblGrid')
    for w in ws:
        grid.append(el('w:gridCol', **{'w:w': w}))
    tbl.append(grid)

    B_TIT = (P['borde_fuerte'],) * 4
    B_ENC = (P['borde_fuerte'],) * 4

    # Fila 1: titulo del anexo, combinado a lo ancho
    tbl.append(fila([celda(titulo, ancho_total, P['titulo_fondo'], negrita=True,
                           color=P['titulo_texto'], gridspan=len(campos),
                           bordes_=B_TIT, margen_lat=40)],
                    alto=340, encabezado_repetible=REPETIR_TITULO))

    # Filas 2 y 3: encabezados. Las columnas agrupadas (dotacion por turno)
    # quedan bajo un encabezado combinado; el resto se fusiona verticalmente.
    hay_grupos = any(c[4] for c in campos)
    if hay_grupos:
        f_a, f_b, i = [], [], 0
        while i < len(campos):
            clave, tit, _, _, grupo = campos[i]
            if grupo:
                j = i
                while j < len(campos) and campos[j][4] == grupo:
                    j += 1
                span = j - i
                f_a.append(celda(grupo, sum(ws[i:j]), P['grupo_fondo'],
                                 negrita=True, color=P['grupo_texto'],
                                 gridspan=span, bordes_=B_ENC, margen_lat=40))
                for k in range(i, j):
                    f_b.append(celda(campos[k][1], ws[k], P['encabezado_fondo'],
                                     negrita=True, color=P['encabezado_texto'],
                                     bordes_=B_ENC, margen_lat=40))
                i = j
            else:
                f_a.append(celda(tit, ws[i], P['encabezado_fondo'], negrita=True,
                                 color=P['encabezado_texto'], vmerge='restart',
                                 bordes_=B_ENC))
                f_b.append(celda('', ws[i], P['encabezado_fondo'],
                                 vmerge='continue', bordes_=B_ENC))
                i += 1
        tbl.append(fila(f_a, encabezado_repetible=REPETIR_ENCABEZADO))
        tbl.append(fila(f_b, encabezado_repetible=REPETIR_ENCABEZADO))
    else:
        tbl.append(fila([celda(c[1], w, P['encabezado_fondo'], negrita=True,
                               color=P['encabezado_texto'], bordes_=B_ENC)
                         for c, w in zip(campos, ws)],
                        encabezado_repetible=REPETIR_ENCABEZADO))

    # Datos
    zebra = P['zebra'] if (USAR_ZEBRA and P['zebra']) else None
    for n, reg in enumerate(registros):
        fondo = zebra if (zebra and n % 2) else P['dato']
        celdas = []
        for (clave, _, _, _, _), w in zip(campos, ws):
            val = reg.get(clave, '')
            izq = clave in ('descripcion', 'observaciones') \
                and orientacion == 'apaisado'
            celdas.append(celda(val, w, fondo,
                                alineacion='left' if izq else 'center',
                                bordes_=(P['borde'],) * 4))
        tbl.append(fila(celdas, entera=FILA_ENTERA))
    return tbl


def tabla_indice(grupos):
    """Indice de anexos: una fila por anexo, agrupadas por familia y con
    link interno al anexo correspondiente."""
    P = PALETAS[PALETA]
    ancho = ANCHO_INDICE
    texto_area = PAGINA['vertical'] - 2 * MARGEN_SECCION

    tbl = el('w:tbl')
    pr = el('w:tblPr')
    pr.append(el('w:tblStyle', **{'w:val': 'a0'}))
    pr.append(el('w:tblW', **{'w:w': ancho, 'w:type': 'dxa'}))
    pr.append(el('w:tblInd', **{'w:w': max(0, (texto_area - ancho) // 2),
                                'w:type': 'dxa'}))
    tb = el('w:tblBorders')
    for lado in ('top', 'left', 'bottom', 'right', 'insideH', 'insideV'):
        tb.append(el(f'w:{lado}', **{'w:val': 'nil'}))
    pr.append(tb)
    pr.append(el('w:tblLayout', **{'w:type': 'fixed'}))
    tbl.append(pr)

    grid = el('w:tblGrid')
    grid.append(el('w:gridCol', **{'w:w': ancho}))
    tbl.append(grid)

    B = (P['borde_fuerte'],) * 4
    tbl.append(fila([celda('ANEXOS', ancho, P['titulo_fondo'], negrita=True,
                           color=P['titulo_texto'], bordes_=B, tam=22)],
                    alto=400))

    familias = []
    for g in grupos:
        if g.get('familia') not in familias:
            familias.append(g.get('familia'))

    for familia in familias:
        dela = [g for g in grupos if g.get('familia') == familia]
        if familia:
            tbl.append(fila([celda(familia, ancho, P['grupo_fondo'],
                                   negrita=True, color=P['grupo_texto'],
                                   bordes_=B, tam=18)], alto=280))
        for g in dela:
            texto = f"ANEXO {g['titulo']}"
            if g.get('detalle'):
                texto += '\n' + g['detalle']
            if not g.get('sin_conteo'):
                n = len(g['regs'])
                texto += f"\n{n} servicio{'s' if n != 1 else ''}"
            tbl.append(fila([celda(
                texto, ancho, P['dato'],
                color=P['titulo_fondo'] if INDICE_CLICKEABLE else None,
                subrayado=INDICE_CLICKEABLE,
                enlace=g['marcador'] if INDICE_CLICKEABLE else None,
                bordes_=(P['borde'],) * 4, tam=20)],
                entera=True))
    return tbl


def parrafo(texto='', *, negrita=False, tam=28, salto_pagina=False,
            centrado=True, marcador=None, bid=0):
    p = el('w:p')
    pr = el('w:pPr')
    if centrado:
        pr.append(el('w:jc', **{'w:val': 'center'}))
    p.append(pr)
    r = el('w:r')
    r_pr = el('w:rPr')
    if negrita:
        r_pr.append(el('w:b'))
        r_pr.append(el('w:bCs'))
    r_pr.append(el('w:sz', **{'w:val': tam}))
    r_pr.append(el('w:szCs', **{'w:val': tam}))
    r.append(r_pr)
    if salto_pagina:
        r.append(el('w:br', **{'w:type': 'page'}))
    if texto:
        t = el('w:t')
        t.set(qn('xml:space'), 'preserve')
        t.text = texto
        r.append(t)
    p.append(r)
    if marcador:
        # el destino del link del indice
        p.insert(1, el('w:bookmarkStart', **{'w:id': str(bid),
                                             'w:name': marcador}))
        p.append(el('w:bookmarkEnd', **{'w:id': str(bid)}))
    return p


def cortar_a_apaisado(doc, marca):
    """Cierra la seccion vertical justo antes de los anexos y deja el resto
    del documento apaisado. El encabezado con los logos se conserva porque
    la referencia viaja dentro del sectPr que se copia."""
    body = doc.element.body
    sect = body.find(W + 'sectPr')
    p = el('w:p')
    ppr = el('w:pPr')
    ppr.append(copy.deepcopy(sect))
    p.append(ppr)
    marca.addprevious(p)

    pg = sect.find(W + 'pgSz')
    ancho, alto = pg.get(qn('w:w')), pg.get(qn('w:h'))
    pg.set(qn('w:w'), alto)
    pg.set(qn('w:h'), ancho)
    pg.set(qn('w:orient'), 'landscape')


NS_WP = '{http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing}'
NS_A = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
NS_R = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'


MC_IGNORABLE = '{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable'


def _volcar_y_escalar(snap, destino, ancho_objetivo_emu):
    """Vuelca el contenido guardado de un encabezado/pie en una parte nueva
    (rehaciendo las relaciones de imagen) y escala el banner al ancho pedido,
    manteniendo la proporcion para no deformar los logos."""
    from docx.opc.constants import RELATIONSHIP_TYPE as RT

    elemento, rels = snap

    # Se clona el elemento RAIZ, no sus hijos: asi viajan tambien las
    # declaraciones de namespace y el mc:Ignorable. Copiando solo los hijos
    # el archivo queda corrupto, porque el wp14 del banner no esta declarado.
    dest_el = copy.deepcopy(elemento)
    destino.part._element = dest_el

    mapa = {}
    for rid, rel in rels:
        if rel.reltype == RT.IMAGE:
            mapa[rid] = destino.part.relate_to(rel.target_part, RT.IMAGE)
    for e in dest_el.iter():
        for attr in (NS_R + 'embed', NS_R + 'link'):
            v = e.get(attr)
            if v in mapa:
                e.set(attr, mapa[v])

    ext = dest_el.find(f'.//{NS_WP}extent')
    if ext is None:
        return 0
    escala = ancho_objetivo_emu / int(ext.get('cx'))
    alto = 0
    for tag in (f'.//{NS_WP}extent', f'.//{NS_A}ext'):
        for e in dest_el.findall(tag):
            if e.get('cx') is None:
                continue
            e.set('cx', str(int(int(e.get('cx')) * escala)))
            e.set('cy', str(int(int(e.get('cy')) * escala)))
            alto = max(alto, int(e.get('cy')))
    return alto


def ajustar_encabezado_apaisado(doc):
    """Le da a la seccion de anexos un encabezado y un pie propios, con el
    banner a lo ancho de la hoja apaisada, y le abre los margenes.

    Ojo: las dos secciones arrancan compartiendo header1.xml/footer1.xml.
    Hay que sacarle la referencia a la seccion apaisada ANTES de pedirle a
    python-docx una parte nueva, si no devuelve la compartida y se termina
    editando el encabezado de las paginas verticales."""
    if len(doc.sections) < 2:
        return
    sec = doc.sections[-1]
    previa = doc.sections[0]
    ancho_emu = PAGINA['apaisado'] * 635   # 1 dxa = 635 EMU

    if BANNER_ANCHO_COMPLETO:
        snaps, compartidas = {}, {}
        for attr in ('header', 'footer'):
            o = getattr(previa, attr)
            snaps[attr] = (copy.deepcopy(o.part.element), list(o.part.rels.items()))
            compartidas[attr] = o.part

        for tag in ('headerReference', 'footerReference'):
            for e in sec._sectPr.findall(W + tag):
                sec._sectPr.remove(e)

        for attr in ('header', 'footer'):
            destino = getattr(sec, attr)
            try:
                destino.is_linked_to_previous = False
                if destino.part is compartidas[attr]:
                    raise RuntimeError('sigue compartiendo la parte')
                _volcar_y_escalar(snaps[attr], destino, ancho_emu)
            except Exception as e:
                print(f"  (no pude escalar el {attr} apaisado: {e})")

    mar = sec._sectPr.find(W + 'pgMar')
    if mar is not None:
        mar.set(qn('w:top'), str(MARGEN_SUP_ANEXO))
        mar.set(qn('w:bottom'), str(MARGEN_INF_ANEXO))


def volver_a_vertical(doc, antes_de):
    """Cierra la seccion apaisada y deja el resto del documento en vertical,
    para que el anexo de imagenes salga en hoja normal."""
    body = doc.element.body
    sect = body.find(W + 'sectPr')
    p = el('w:p')
    ppr = el('w:pPr')
    ppr.append(copy.deepcopy(sect))
    p.append(ppr)
    antes_de.addprevious(p)

    pg = sect.find(W + 'pgSz')
    ancho, alto = pg.get(qn('w:w')), pg.get(qn('w:h'))
    pg.set(qn('w:w'), alto)
    pg.set(qn('w:h'), ancho)
    if pg.get(qn('w:orient')):
        del pg.attrib[qn('w:orient')]
    mar = sect.find(W + 'pgMar')
    if mar is not None:
        mar.set(qn('w:top'), str(MARGEN_SECCION))
        mar.set(qn('w:bottom'), str(MARGEN_SECCION))


def anexar_documento(doc, ruta, antes_de):
    """Copia el cuerpo de otro .docx al final del documento, rehaciendo las
    relaciones para que las imagenes y los links no se pierdan.
    Devuelve cuantos elementos copio."""
    from docx.opc.constants import RELATIONSHIP_TYPE as RT

    origen = Document(ruta)
    mapa = {}
    for rid, rel in origen.part.rels.items():
        if rel.is_external:
            mapa[rid] = doc.part.relate_to(rel.target_ref, rel.reltype,
                                           is_external=True)
        elif rel.reltype == RT.IMAGE:
            # Hay que REINSERTAR los bytes, no relacionar la parte de origen:
            # los dos documentos salen de la misma plantilla y tienen partes
            # con el mismo nombre (word/media/image1.png). Relacionarlas
            # directamente hace colisionar los nombres y el .docx queda roto.
            nuevo_rid, _ = doc.part.get_or_add_image(
                io.BytesIO(rel.target_part.blob))
            mapa[rid] = nuevo_rid

    copiados, salteados = 0, 0
    for hijo in origen.element.body:
        if hijo.tag == W + 'sectPr':
            continue
        if MARCA_INSTRUCCIONES in ''.join(
                n.text or '' for n in hijo.iter() if n.tag == W + 't'):
            salteados += 1
            continue
        nuevo = copy.deepcopy(hijo)
        for e in nuevo.iter():
            for attr in (NS_R + 'embed', NS_R + 'link', NS_R + 'id'):
                v = e.get(attr)
                if v in mapa:
                    e.set(attr, mapa[v])
        antes_de.addprevious(nuevo)
        copiados += 1
    return copiados, salteados


def quitar_comentarios(doc):
    """El documento final va a SADE: no deberia llevar notas internas."""
    body = doc.element.body
    for tag in ('commentRangeStart', 'commentRangeEnd'):
        for e in body.findall(f'.//{W}{tag}'):
            e.getparent().remove(e)
    for ref in body.findall(f'.//{W}commentReference'):
        run = ref.getparent()
        if run.tag == W + 'r':
            run.getparent().remove(run)
        else:
            run.remove(ref)
    try:
        for part in doc.part.package.iter_parts():
            if str(part.partname).endswith('/comments.xml'):
                for hijo in list(part.element):
                    part.element.remove(hijo)
    except Exception:
        pass


def pdf_bloqueado(pdf_path):
    """Word falla con un error opaco si el PDF destino esta abierto en un
    visor. Conviene detectarlo antes para poder avisar en criollo."""
    if not Path(pdf_path).exists():
        return False
    try:
        with open(pdf_path, 'r+b'):
            return False
    except OSError:
        return True


def numero_de_orden(doc):
    """Saca el numero del titulo de la plantilla: de
    'Orden de Servicio N° 2006-2026-O' devuelve '2006-2026-O'."""
    for p in doc.element.body.findall(W + 'p'):
        txt = ''.join(n.text or '' for n in p.iter() if n.tag == W + 't').strip()
        if txt.lower().startswith('orden de servicio'):
            m = re.search(r'N[°ºo]\s*([\w./-]+)', txt)
            if m:
                return m.group(1).strip('.-')
            break
    return None


def buscar_planilla(ruta=None):
    """Si no se pasa --xlsx, busca en Descargas la planilla mas reciente que
    coincida con el patron, ignorando el '(1)' que agrega el navegador."""
    if ruta:
        p = Path(ruta).expanduser()
        if not p.exists():
            sys.exit(f"No encuentro la planilla: {p}")
        return p

    CARPETA_PLANILLA.mkdir(exist_ok=True)
    for carpeta, donde in ((CARPETA_PLANILLA, 'la carpeta planilla'),
                           (CARPETA_DESCARGAS, 'Descargas')):
        candidatos = sorted(
            (c for c in carpeta.glob(PATRON_XLSX)
             if not c.name.startswith('~$')),
            key=lambda x: x.stat().st_mtime, reverse=True)
        if not candidatos:
            continue
        elegida = candidatos[0]
        if len(candidatos) > 1:
            print(f"\n  Hay {len(candidatos)} planillas en {donde}; "
                  f"uso la mas reciente:")
            for c in candidatos:
                marca = '  <-- esta' if c is elegida else ''
                fecha = datetime.datetime.fromtimestamp(c.stat().st_mtime)
                print(f"     {fecha:%d/%m %H:%M}  {c.name}{marca}")
            print("     (si no es la correcta, borra las viejas o usa --xlsx)")
        else:
            print(f"\n  Planilla: {elegida.name}   (en {donde})")
        return elegida

    sys.exit(
        f"\nNo encontre la planilla.\n\n"
        f"Baja el libro desde Excel Online y dejalo en esta carpeta:\n\n"
        f"    {CARPETA_PLANILLA}\n\n"
        f"El nombre puede tener sufijos: 'SISTEMA DE PREORDEN (1).xlsx' "
        f"se toma igual.\n"
        f"Tambien se busca en {CARPETA_DESCARGAS}.\n")


def buscar_soffice():
    """LibreOffice, para convertir a PDF donde no hay Word por COM (Mac)."""
    import shutil as sh
    encontrado = sh.which('soffice') or sh.which('libreoffice')
    if encontrado:
        return encontrado
    candidatos = [
        '/Applications/LibreOffice.app/Contents/MacOS/soffice',
        r'C:\Program Files\LibreOffice\program\soffice.exe',
        r'C:\Program Files (x86)\LibreOffice\program\soffice.exe',
        '/usr/bin/soffice', '/usr/local/bin/soffice',
        '/opt/homebrew/bin/soffice',
    ]
    return next((c for c in candidatos if Path(c).exists()), None)


def _pdf_con_word(docx_path, pdf_path):
    import win32com.client as win32
    word = win32.DispatchEx("Word.Application")
    word.Visible = False
    try:
        d = word.Documents.Open(str(docx_path), ReadOnly=True)
        d.SaveAs(str(pdf_path), FileFormat=17)   # 17 = wdFormatPDF
        d.Close(False)
    finally:
        word.Quit()


def _pdf_con_libreoffice(docx_path, pdf_path, soffice):
    import subprocess, shutil as sh
    salida = Path(pdf_path).parent
    subprocess.run([soffice, '--headless', '--convert-to', 'pdf',
                    '--outdir', str(salida), str(docx_path)],
                   check=True, capture_output=True, timeout=300)
    # LibreOffice nombra el PDF como el .docx: hay que renombrarlo
    generado = salida / (Path(docx_path).stem + '.pdf')
    if generado != Path(pdf_path):
        sh.move(str(generado), str(pdf_path))


def exportar_pdf(docx_path, pdf_path):
    """Windows usa Word; en Mac o Linux cae a LibreOffice."""
    if sys.platform == 'win32':
        try:
            _pdf_con_word(docx_path, pdf_path)
            return pdf_path
        except ImportError:
            pass
    soffice = buscar_soffice()
    if not soffice:
        raise RuntimeError(
            "no hay con qué convertir a PDF. En Windows hace falta Word "
            "(con pywin32 instalado); en Mac, LibreOffice "
            "(brew install --cask libreoffice). El .docx se generó igual.")
    _pdf_con_libreoffice(docx_path, pdf_path, soffice)
    return pdf_path


# ==========================================================================
# ARMADO DEL DOCUMENTO
# ==========================================================================

def ordenar(regs, orden_bases=None):
    if ORDEN_SERVICIOS == 'planilla':
        return sorted(regs, key=lambda s: s['fila'])
    pos = (lambda s: orden_bases.index(s['comuna'])) if orden_bases else (lambda s: 0)
    if ORDEN_SERVICIOS == 'cobertura':
        k = lambda s: (pos(s), clave_cobertura(s['dia']),
                       clave_turno(s['turno']), clave_hora(s['hora']), s['fila'])
    elif ORDEN_SERVICIOS == 'base-turno-dia':
        k = lambda s: (pos(s), clave_turno(s['turno']), clave_dia(s['dia']),
                       clave_hora(s['hora']), s['fila'])
    elif ORDEN_SERVICIOS == 'base-dia-hora':
        k = lambda s: (pos(s), clave_dia(s['dia']), clave_hora(s['hora']), s['fila'])
    elif ORDEN_SERVICIOS == 'base-hora':
        k = lambda s: (pos(s), clave_hora(s['hora']), s['fila'])
    elif ORDEN_SERVICIOS == 'dia-hora':
        k = lambda s: (clave_dia(s['dia']), clave_hora(s['hora']), s['fila'])
    else:
        k = lambda s: s['fila']
    return sorted(regs, key=k)


def agrupar(servicios, ambito, solo=None):
    """Devuelve los anexos en el orden en que van al documento: primero las
    bases operativas, despues las zonas comunales."""
    grupos = []

    def coincide(titulo):
        return not solo or solo.upper() in titulo.upper()

    if ambito in ('completa', 'bases'):
        de_bases = [s for s in servicios if s['comuna'] is None]
        vistas = [b for b in ORDEN_BASES_OPERATIVAS
                  if any(s['base_cruda'] == b for s in de_bases)]
        otras = sorted({s['base_cruda'] for s in de_bases
                        if s['base_cruda'] not in vistas})
        for base in vistas + otras:
            titulo = f"BASE {base}"
            if not coincide(titulo):
                continue
            regs = [s for s in de_bases if s['base_cruda'] == base]
            grupos.append(dict(familia='BASES OPERATIVAS', titulo=titulo,
                               detalle='', regs=ordenar(regs)))

    if ambito in ('completa', 'comunas'):
        de_comunas = [s for s in servicios if s['comuna'] is not None]
        for zona, comunas in ZONAS.items():
            titulo = f"ZONA COMUNAL {zona}"
            if not coincide(titulo):
                continue
            regs = [s for s in de_comunas if s['comuna'] in comunas]
            lista = ', '.join(str(c) for c in comunas[:-1])
            grupos.append(dict(familia='ZONAS COMUNALES', titulo=titulo,
                               detalle=f"(Comunas {lista} y {comunas[-1]})",
                               regs=ordenar(regs, comunas)))

    if OMITIR_ANEXOS_VACIOS and any(g['regs'] for g in grupos):
        grupos = [g for g in grupos if g['regs']]

    for i, g in enumerate(grupos, 1):
        g['marcador'] = f'anexo{i}'
    return grupos


def generar(xlsx, ambito, finde, plantilla, salida, orientacion, solo=None):
    hoja = HOJA_CARGA[bool(finde)]
    campos = CAMPOS_FINDE if finde else CAMPOS_SEMANA
    servicios, sin_base = leer_servicios(xlsx, finde)

    en_comuna = [s for s in servicios if s['comuna'] is not None]
    if finde and FINDE_SOLO_BASES:
        ambito = 'bases'
        if en_comuna:
            print(f"\n  !! {len(en_comuna)} servicio(s) de FINDE con base de "
                  f"despliegue comunal. Las comunas no trabajan fines de\n"
                  f"     semana, asi que quedan afuera. Revisar la carga:")
            for s in en_comuna[:10]:
                print(f"     fila {s['fila']}: {s['id']} - {s['base_cruda']}")

    grupos = agrupar(servicios, ambito, solo)
    if not grupos:
        sys.exit(f"No hay ningun anexo que coincida con {solo!r}")

    imagenes = ANEXO_IMAGENES_FINDE if finde else ANEXO_IMAGENES_SEMANA
    if INCLUIR_ANEXO_IMAGENES and imagenes.exists() and not solo:
        grupos.append(dict(familia=None, titulo='DE IMÁGENES', detalle='',
                           regs=[], marcador='anexoimagenes',
                           sin_conteo=True, imagenes=imagenes))
    else:
        imagenes = None

    doc = Document(plantilla)
    body = doc.element.body

    def buscar(nombre):
        for p in body.findall(W + 'p'):
            txt = ''.join(n.text or '' for n in p.iter() if n.tag == W + 't')
            if nombre in txt:
                return p
        return None

    marca = buscar('{{ANEXOS}}')
    if marca is None:
        sys.exit(f"La plantilla {plantilla.name} no tiene el marcador {{{{ANEXOS}}}}")

    indice = buscar('{{INDICE}}')
    if indice is not None:
        indice.addprevious(tabla_indice(grupos))
        body.remove(indice)

    if orientacion == 'apaisado':
        cortar_a_apaisado(doc, marca)
        ajustar_encabezado_apaisado(doc)

    copiados = 0
    for i, g in enumerate(grupos):
        if g.get('imagenes') and orientacion == 'apaisado' \
                and ANEXO_IMAGENES_ORIENTACION == 'vertical':
            volver_a_vertical(doc, marca)
        elif i or orientacion != 'apaisado':
            # con salto de seccion la primera pagina ya viene sola
            marca.addprevious(parrafo(salto_pagina=True))

        marca.addprevious(parrafo(f"ANEXO {g['titulo']}", negrita=True, tam=28,
                                  marcador=g['marcador'], bid=100 + i))
        marca.addprevious(parrafo())

        if g.get('imagenes'):
            copiados, salteados = anexar_documento(doc, g['imagenes'], marca)
        elif g['regs']:
            marca.addprevious(tabla_anexo(g['titulo'], g['regs'], campos,
                                          orientacion))
        else:
            marca.addprevious(parrafo('(sin servicios validados para este anexo)',
                                      tam=20))
        marca.addprevious(parrafo())

    body.remove(marca)
    if LIMPIAR_COMENTARIOS:
        quitar_comentarios(doc)

    salida.parent.mkdir(parents=True, exist_ok=True)
    try:
        doc.save(salida)
    except PermissionError:
        sys.exit(
            f"\nNo puedo escribir {salida.name}: el archivo esta abierto.\n"
            f"Cerralo en Word y volve a correr el generador.\n"
            f"   ({salida})\n")

    print(f"\n=== {hoja} / {ambito} / {orientacion} / orden {ORDEN_SERVICIOS} ===")
    total, familia_previa = 0, None
    for g in grupos:
        if g.get('familia') and g['familia'] != familia_previa:
            print(f"  -- {g['familia']}")
            familia_previa = g['familia']
        if g.get('imagenes'):
            print(f"     ANEXO {g['titulo']:26s} {copiados:4d} elementos "
                  f"desde {g['imagenes'].name}")
            continue
        print(f"     {g['titulo']:32s} {len(g['regs']):4d} servicios")
        total += len(g['regs'])
    print(f"  {'TOTAL':37s} {total:4d}")
    if INCLUIR_ANEXO_IMAGENES and not imagenes and not solo:
        falta = ANEXO_IMAGENES_FINDE if finde else ANEXO_IMAGENES_SEMANA
        print(f"  (sin anexo de imagenes: no existe {falta.name})")

    incluidos = [s for g in grupos for s in g['regs']]
    sin_ag = [s for s in incluidos if not s['total_ag']]
    sin_turno = [s for s in incluidos if not s['turno']]
    largas = sorted([s for s in incluidos if s['largo_desc'] > 400],
                    key=lambda s: -s['largo_desc'])
    if sin_base:
        print(f"\n  !! {len(sin_base)} servicio(s) SIN BASE, afuera del documento:")
        for s in sin_base:
            print(f"     fila {s['fila']}: {s['id']}")
    if sin_turno:
        print(f"  !! {len(sin_turno)} sin TURNO cargado")
    num = [s for s in incluidos if s.get('hora_numerica')]
    if num:
        print(f"  !! {len(num)} servicio(s) con la HORA cargada como hora de "
              f"Excel y no como texto.\n"
              f"     Salen bien (HH:MM) pero pierden el rango. Conviene "
              f"escribirlas como '8 A 19HS':")
        for s in num:
            print(f"     fila {s['fila']}: {s['id']} - {s['base_cruda']} -> {s['hora']}")
    turnos = sorted({s['turno'] for s in incluidos if s['turno']})
    if len(turnos) > 6:
        print(f"  !! {len(turnos)} formas distintas de escribir el TURNO "
              f"(conviene unificar con una lista desplegable en la planilla):")
        print("     " + ' | '.join(turnos))
    if sin_ag:
        print(f"  !! {len(sin_ag)} sin dotacion en ningun turno")
    con_obs = [s for s in incluidos if s['observaciones']]
    print(f"  .. {len(con_obs)} de {len(incluidos)} tienen OBSERVACIONES cargadas")
    recortadas = [s for s in incluidos if s.get('recortada')]
    if recortadas:
        print(f"  !! {len(recortadas)} DESCRIPCION(es) recortadas a "
              f"{MAX_DESCRIPCION} caracteres; conviene acortarlas en la planilla:")
        for s in recortadas:
            print(f"     fila {s['fila']}: {s['id']} - {s['largo_desc']} caracteres")
    elif largas:
        print(f"  !! {len(largas)} con FUNCION muy larga (max "
              f"{largas[0]['largo_desc']} caracteres, fila {largas[0]['fila']})")

    print(f"\n  -> {salida}")

    if EXPORTAR_PDF:
        numero = numero_de_orden(doc)
        if numero:
            etiqueta = numero
        else:
            etiqueta = {'completa': 'Semanal', 'comunas': 'Comunas',
                        'bases': 'Bases'}[ambito]
            if finde:
                etiqueta = 'Finde'
            etiqueta += f"_{NUMERO_ORDEN}_{ANIO_ORDEN}"
            print(f"  (la plantilla no tiene N° de orden en el titulo; "
                  f"uso {etiqueta})")
        if solo:
            etiqueta += '_' + re.sub(r'\W+', '', solo)
        pdf = salida.parent / f"ODS_{etiqueta}.pdf"
        if pdf_bloqueado(pdf):
            print(f"\n  !! El Word se generó bien, pero NO pude actualizar el PDF:\n"
                  f"     {pdf.name} está abierto en otro programa.\n"
                  f"     Cerralo y volvé a correr el generador.")
        else:
            try:
                exportar_pdf(salida.resolve(), pdf.resolve())
                print(f"  -> {pdf}")
            except Exception as e:
                print(f"  (no pude exportar el PDF: {e})")
    return salida


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default=None,
                    help='planilla a usar; por defecto la mas reciente de Descargas')
    ap.add_argument('--ambito', choices=['completa', 'comunas', 'bases'],
                    default='completa',
                    help='completa = bases + comunas en un solo documento')
    ap.add_argument('--finde', action='store_true',
                    help='generar SOLO la orden de fin de semana')
    ap.add_argument('--semana', action='store_true',
                    help='generar SOLO la orden semanal')
    ap.add_argument('--vertical', action='store_true',
                    help='forzar anexos verticales')
    ap.add_argument('--zona', default=None,
                    help='generar un solo anexo (CENTRO, NORTE, SUR o el nombre de la base)')
    ap.add_argument('--paleta', choices=sorted(PALETAS), default=None,
                    help='colores de la tabla')
    ap.add_argument('--max-desc', type=int, default=None,
                    help='recorte de la columna DESCRIPCION (0 = sin recorte)')
    ap.add_argument('--tam', type=int, default=None,
                    help='tamano de letra en puntos (9 por defecto)')
    ap.add_argument('--salida', default=None)
    a = ap.parse_args()

    if a.paleta:
        globals()['PALETA'] = a.paleta
    if a.max_desc is not None:
        globals()['MAX_DESCRIPCION'] = a.max_desc or None
    if a.tam:
        globals()['TAM'] = a.tam * 2

    xlsx = buscar_planilla(a.xlsx)
    orientacion = 'vertical' if a.vertical else ORIENTACION

    # Sin flags se generan las dos ordenes: es lo que hace falta cada semana
    # y olvidarse del --finde era el error facil de cometer.
    if a.finde:
        ordenes = [True]
    elif a.semana:
        ordenes = [False]
    elif a.salida or a.zona:
        ordenes = [False]     # apuntan a un documento puntual
    else:
        ordenes = [False, True]

    for finde in ordenes:
        plantilla = PLANTILLA_FINDE if (finde and PLANTILLA_FINDE.exists()) \
            else PLANTILLA_SEMANA
        if not plantilla.exists():
            sys.exit(f"No encuentro la plantilla: {plantilla}")
        if finde and not PLANTILLA_FINDE.exists():
            print("  AVISO: no existe plantilla_OS_FINDE.docx; uso la semanal.")

        nombre = "Orden de servicio " + ('FINDE' if finde else 'SEMANA')
        if a.ambito != 'completa':
            nombre += f" ({a.ambito})"
        if a.zona:
            nombre += f" - {a.zona.upper()}"
        if orientacion == 'vertical':
            nombre += " (vertical)"
        salida = Path(a.salida) if a.salida else DIR_SALIDA / f"{nombre}.docx"

        generar(xlsx, a.ambito, finde, plantilla, salida, orientacion, a.zona)


if __name__ == '__main__':
    main()
